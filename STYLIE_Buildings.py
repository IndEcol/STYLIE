# -*- coding: utf-8 -*-
"""
Created on April 12, 2024

@author: spauliuk

This script loads previously compiled results and then compiles selected results 
into different visualisations of the energy service cascade.

Works together with IAMC data template and control workbook
STYLIE_ESC_Configure.xlsx

Documentation and how to available in https://github.com/IndEcol/STYLIE
"""

import os
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D
import numpy as np


def get_esc_data_from_pandas(ps,selectI,selectR,cscenss):
    # return numpay array with esc data for given indicator, region, and scenario list
    pst     = ps[ps['Indicator'].isin([selectI]) & ps['Region'].isin(selectR) & ps['Scenario'].isin(cscenss)] # Select the specified data and transpose them for plotting
    unit    = pst.iloc[0]['Unit']
    Data    = pst.drop(['Scenario','Indicator', 'Region', 'Unit'], axis=1).values
    return Data, unit

plt.style.use('default') # set all plotting parameters to their default values

# Definitions/Specifications
CF            = openpyxl.load_workbook('STYLIE_ESC_Configure.xlsx')
CS            = CF['Cover'].cell(4,4).value

scen = [] # list of target scenarios


r = 1
# Move to parameter list:
while True:
    if CF[CS].cell(r,1).value == 'Define ESC plot':
        break    
    r += 1
r += 1

ctitles = []
ctypes  = []
cregs   = []
cscens  = []
colors  = [] # List with color strings

while True:
    if CF[CS].cell(r,2).value is None:
        break    
    ctitles.append(CF[CS].cell(r,2).value)
    ctypes.append(CF[CS].cell(r,3).value)
    cregs.append(CF[CS].cell(r,4).value)
    cscens.append(CF[CS].cell(r,5).value)
    colors.append(CF[CS].cell(r,11).value)
    r += 1


# open data file with results
ps = pd.read_excel('IAMC_Template_RECCv2.5_SampleData.xlsx', sheet_name='Buildings', index_col=0) # plot sheet

# determine ESC indicators and plot ESC cascades
# find ESC parameters in Res array: outer index: Ar, middle index: ti, inner indicator: scen

for c in range(0,len(ctitles)):
    if ctypes[c] == 'version_2_blds': # Energy service cascade_GHG for buildings
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = [cregs[c]]
        
        # Define data container
        esc_data = np.zeros((10,46,nocs)) # 10 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Population:
        Data_pop = np.concatenate([get_esc_data_from_pandas(ps,'Population',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # GHG:
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'Emissions|CO2|Energy|Demand|Residential',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_ghg2 = np.concatenate([get_esc_data_from_pandas(ps,'Emissions|CO2|Energy|Demand|Commercial',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'Emissions|CO2|Energy|Supply',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_ghg4 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, primary material production',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Final energy:
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Final Energy|Residential and Commercial',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Stock:
        Data_rebx = np.concatenate([get_esc_data_from_pandas(ps,'Energy Service|Residential|Floor Space',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_nrbx = np.concatenate([get_esc_data_from_pandas(ps,'Energy Service|Commercial|Floor Space',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Inflow
        Data_rebf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all res. building types together',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_nrbf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all nonres. building types together',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)])        
        # Final material consumption:
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Material footprint / RMI
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_pop).transpose()
        esc_data[1,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_edx).transpose()
        esc_data[2,:,:] = (Data_edx / (Data_rebx + Data_nrbx)).transpose() 
        esc_data[3,:,:] = ((Data_rebx + Data_nrbx)/Data_pop).transpose()
        esc_data[4,:,:] = ((Data_matm) / (Data_rebf + Data_nrbf)).transpose()
        esc_data[5,:,:] = (Data_ghg4 / Data_matm).transpose()
        esc_data[6,:,:] = (Data_ghg4 / Data_pop).transpose()
        esc_data[7,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[8,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_pop).transpose()
        esc_data[9,:,:] = ((Data_rebf + Data_nrbf)/(Data_rebx + Data_nrbx)).transpose()
        
        # Define maximal GHG/cap
        maxGHG = np.max(esc_data[[0,6],1::,:])
        
        # Define colors
        ccol = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple        
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=8 , figsize=(24, 3))        
        fig.suptitle('Energy and material service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", ccol)
        
        axs[0].plot(np.arange(2016,2061), esc_data[0,1::,:],   linewidth = 3)
        plta = Line2D(np.arange(2016,2061), esc_data[0,1::,:], linewidth = 3)
        ProxyHandlesList.append(plta) # create proxy artist for legend    
        axs[0].set_title('Energy-GHG per capita     = \n (Scope 1 + 2 emissions)      ', weight='bold')
        axs[0].set_ylabel('t CO2-eq/yr', fontsize = 12)
        #axs[0].set_facecolor((221/255, 235/255, 247/255))
        axs[0].set_facecolor((197/255, 221/255, 241/255))
        axs[0].set_ylim(bottom=0)
        axs[0].set_ylim(top=1.05 * maxGHG)
        
        axs[1].plot(np.arange(2016,2061), esc_data[1,1::,:] * 1e6, linewidth = 2.0)  
        axs[1].set_title('GHG per final energy     *')
        axs[1].set_ylabel('g CO2-eq/MJ', fontsize = 12)
        axs[1].set_facecolor((238/255, 245/255, 252/255))
        axs[1].set_ylim(bottom=0)
        
        axs[2].plot(np.arange(2016,2061), esc_data[2,1::,:], linewidth = 2.0) 
        axs[2].set_title('Final energy per stock     *')
        axs[2].set_ylabel('MJ/(m²·yr)', fontsize = 12)
        axs[2].set_facecolor((238/255, 245/255, 252/255))
        axs[2].set_ylim(bottom=0)
        
        axs[3].plot(np.arange(2016,2061), esc_data[3,1::,:], linewidth = 3.0)  
        axs[3].set_title(' <-- | --> \n Stock per capita', weight='bold')
        axs[3].set_ylabel('m²', fontsize = 12)
        axs[3].set_facecolor((237/255, 226/255, 246/255))  
        axs[3].set_ylim(bottom=0)
        
        axs[4].plot(np.arange(2016,2061), esc_data[9,1::,:], linewidth = 2.0)   
        axs[4].set_title('*     inflow per stock')        
        axs[4].set_ylabel('1/yr', fontsize = 12)
        axs[4].set_facecolor((253/255, 239/255, 231/255))            
        axs[4].set_ylim(bottom=0)
        
        axs[5].plot(np.arange(2016,2061), esc_data[4,1::,:] * 1e3, linewidth = 2.0)   
        axs[5].set_title('*     material intensity of inflow')        
        axs[5].set_ylabel('kg/m²', fontsize = 12)
        axs[5].set_facecolor((253/255, 239/255, 231/255))            
        axs[5].set_ylim(bottom=0)
        
        axs[6].plot(np.arange(2016,2061), esc_data[5,1::,:], linewidth = 2.0) 
        axs[6].set_title('*     material-GHG per material')
        axs[6].set_ylabel('t CO2-eq/t', fontsize = 12)
        axs[6].set_facecolor((253/255, 239/255, 231/255))            
        axs[6].set_ylim(bottom=0)
        
        axs[7].plot(np.arange(2016,2061), esc_data[6,1::,:], linewidth = 3.0) 
        axs[7].set_title('=     material-GHG per capita \n(Scope 3 emissions)', weight='bold')
        axs[7].set_ylabel('t CO2-eq/yr', fontsize = 12)
        #axs[7].set_facecolor((252/255, 228/255, 214/255))         
        axs[7].set_facecolor((249/255, 203/255, 177/255))   
        axs[7].set_ylim(bottom=0)
        axs[7].set_ylim(top=1.05 * maxGHG)
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=5, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(title + '_' + selectR[0] + '.png', dpi=150, bbox_inches='tight')


    if ctypes[c] == 'version_3_blds': # Energy service cascade_RMI for buildings
        # get scenario list and length
        if cscens[c] == 'All':
            cscenss = scen
        else:
            cscenss = cscens[c].split(';')
        nocs = len(cscenss)
        selectR = [cregs[c]]
        
        # Define data container
        esc_data = np.zeros((10,46,nocs)) # 10 decoupling indices, 46 years, nocs scenarios
        
        # EXTRACT data and convert to ESC data array
        # Population:
        Data_pop = np.concatenate([get_esc_data_from_pandas(ps,'Population',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # GHG:
        Data_ghg1 = np.concatenate([get_esc_data_from_pandas(ps,'Emissions|CO2|Energy|Demand|Residential',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_ghg2 = np.concatenate([get_esc_data_from_pandas(ps,'Emissions|CO2|Energy|Demand|Commercial',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_ghg3 = np.concatenate([get_esc_data_from_pandas(ps,'Emissions|CO2|Energy|Supply',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_ghg4 = np.concatenate([get_esc_data_from_pandas(ps,'GHG emissions, primary material production',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Final energy:
        Data_edx = np.concatenate([get_esc_data_from_pandas(ps,'Final Energy|Residential and Commercial',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Stock:
        Data_rebx = np.concatenate([get_esc_data_from_pandas(ps,'Energy Service|Residential|Floor Space',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_nrbx = np.concatenate([get_esc_data_from_pandas(ps,'Energy Service|Commercial|Floor Space',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Inflow
        Data_rebf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all res. building types together',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_nrbf = np.concatenate([get_esc_data_from_pandas(ps,'final consumption (use phase inflow), all nonres. building types together',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)])        
        # Final material consumption:
        Data_matm = np.concatenate([get_esc_data_from_pandas(ps,'Final consumption of materials',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        # Material footprint / RMI
        Data_maf1 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, metal ores, system-wide',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_maf2 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, non-metallic minerals, system-wide',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 
        Data_maf3 = np.concatenate([get_esc_data_from_pandas(ps,'Material footprint, biomass (dry weight), system-wide',selectR,[cscenss[mscen]])[0] for mscen in range(0,nocs)]) 

        esc_data[0,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_pop).transpose()
        esc_data[1,:,:] = ((Data_ghg1 + Data_ghg2 + Data_ghg3)/Data_edx).transpose()
        esc_data[2,:,:] = (Data_edx / (Data_rebx + Data_nrbx)).transpose() 
        esc_data[3,:,:] = ((Data_rebx + Data_nrbx)/Data_pop).transpose()
        esc_data[4,:,:] = ((Data_matm) / (Data_rebf + Data_nrbf)).transpose()
        esc_data[5,:,:] = (Data_ghg4 / Data_matm).transpose()
        esc_data[6,:,:] = (Data_ghg4 / Data_pop).transpose()
        esc_data[7,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_matm).transpose()
        esc_data[8,:,:] = ((Data_maf1 + Data_maf2 + Data_maf3)/Data_pop).transpose()
        esc_data[9,:,:] = ((Data_rebf + Data_nrbf)/(Data_rebx + Data_nrbx)).transpose()        
                
        # Define colors
        cc = np.array([[128,128,128,255],[48,84,150,255],[198,89,17,255],[142,105,0,255],[112,48,160,255]])/255 # grey, blue, red, brown, purple
        
        # Plot results
        fig, axs = plt.subplots(nrows=1, ncols=5 , figsize=(16, 3))        
        fig.suptitle('Energy and material service cascade, ' + cregs[c],fontsize=18)
        ProxyHandlesList = []   # For legend 
        
        plt.rcParams["axes.prop_cycle"] = plt.cycler("color", cc)
                
        axs[0].plot(np.arange(2016,2061), esc_data[3,1::,:], linewidth = 3.0)  
        axs[0].set_title('Stock per capita', weight='bold')
        axs[0].set_ylabel('m²', fontsize = 12)
        axs[0].set_facecolor((237/255, 226/255, 246/255))  
        axs[0].set_ylim(bottom=0)
        
        axs[1].plot(np.arange(2016,2061), esc_data[9,1::,:], linewidth = 2.0)   
        axs[1].set_title('*     inflow per stock')        
        axs[1].set_ylabel('1/yr', fontsize = 12)
        axs[1].set_facecolor((253/255, 239/255, 231/255))            
        axs[1].set_ylim(bottom=0)

        axs[2].plot(np.arange(2016,2061), esc_data[4,1::,:] * 1e3, linewidth = 2.0)   
        axs[2].set_title('*     material intensity of inflow')        
        axs[2].set_ylabel('kg/m²', fontsize = 12)
        axs[2].set_facecolor((253/255, 239/255, 231/255))            
        axs[2].set_ylim(bottom=0)        
        
        axs[3].plot(np.arange(2016,2061), esc_data[7,1::,:], linewidth = 2.0) 
        axs[3].set_title('*     RMI per material')
        axs[3].set_ylabel('t/t', fontsize = 12)
        axs[3].set_facecolor((253/255, 239/255, 231/255))            
        axs[3].set_ylim(bottom=0)
        
        axs[4].plot(np.arange(2016,2061), esc_data[8,1::,:], linewidth = 3.0) 
        axs[4].set_title('=     RMI per capita', weight='bold')
        axs[4].set_ylabel('t/yr', fontsize = 12)
        #axs[4].set_facecolor((252/255, 228/255, 214/255))         
        axs[4].set_facecolor((249/255, 203/255, 177/255))         
        axs[4].set_ylim(bottom=0)
        
        Labels = cscenss
        
        fig.legend(Labels, shadow = False, prop={'size':14},ncol=5, loc = 'upper center',bbox_to_anchor=(0.5, -0.02)) 
        plt.tight_layout()
        plt.show()
        title = ctitles[c]
        fig.savefig(title + '_' + selectR[0] + '.png', dpi=150, bbox_inches='tight')
       
             
#
#
#
# The end.
#
#
#